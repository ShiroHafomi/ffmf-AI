"""Tests for Excel upload / preview / export endpoints.

All file content is generated programmatically — no fixture files on disk.
No database dependency: the Excel routes are temp-processing only.
"""

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
import openpyxl
from openpyxl import Workbook

import main


# ──────────────────────── Test helpers ────────────────────────────────
def _empty_wb_bytes(sheet_name="Sheet1"):
    """A workbook with only a header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Date", "Amount", "Category", "Description"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _sample_wb_bytes(rows=None, sheet_name="Sheet1"):
    """Build a workbook with header + data rows."""
    if rows is None:
        rows = [
            ("2026-01-15", "1500000", "Groceries", "Monthly shopping"),
            ("2026-02-20", "2000000", "Utilities", ""),
        ]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Date", "Amount", "Category", "Description"])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _multi_sheet_wb():
    """Workbook with two sheets."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Expenses"
    ws1.append(["Date", "Amount"])
    ws1.append(["2026-01-01", 1000])
    ws1.append(["2026-02-01", 2000])

    ws2 = wb.create_sheet("Budgets")
    ws2.append(["Category", "Budget"])
    ws2.append(["Groceries", 5000000])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _vn_headers_wb():
    """Workbook with Vietnamese column headers."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Ngay", "So tien", "Danh muc", "Ghi chu"])
    ws.append(["2026-03-01", 3000000, "Food", "Lunch"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _excel_serial_dates_wb():
    """Dates stored as Excel serial numbers (int)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Amount"])
    # 2026-01-15 = Excel serial ~46000
    ws.append([46054, 1500000])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _upload(client, filename: str, content: bytes):
    """Helper: upload a file and return the response."""
    return client.post(
        "/api/excel/upload",
        files={"file": (filename, BytesIO(content))},
    )


def _preview(client, filename: str, content: bytes):
    return client.post(
        "/api/excel/preview",
        files={"file": (filename, content)},
    )


@pytest.fixture
def client():
    """Fixture: a ``fastapi.testclient.TestClient`` against ``main.app``."""
    return TestClient(main.app)


# ──────────────────────── Upload tests ─────────────────────────────────
class TestUploadValid:
    """Happy-path uploads with well-formed .xlsx / .csv."""

    def test_upload_xlsx_with_data(self, client):
        content = _sample_wb_bytes()
        resp = _upload(client, "expenses.xlsx", content)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["valid_rows"] == 2
        assert body["total_rows"] == 2
        assert body["skipped_count"] == 0
        assert len(body["data"]) == 2
        assert body["data"][0]["date"] == "2026-01-15"
        assert body["data"][0]["amount"] == 1500000.0
        assert body["data"][0]["category"] == "Groceries"
        assert "headers" in body
        assert "normalised_headers" in body

    def test_upload_csv(self, client):
        csv_content = (
            "Date,Amount,Category,Description\r\n"
            "2026-03-01,500000,Transport,Bus pass\r\n"
        )
        resp = _upload(client, "data.csv", csv_content.encode("utf-8"))
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["valid_rows"] == 1

    def test_upload_vietnamese_headers(self, client):
        content = _vn_headers_wb()
        resp = _upload(client, "expenses.xlsx", content)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["valid_rows"] == 1
        assert body["data"][0]["date"] == "2026-03-01"
        assert body["data"][0]["amount"] == 3000000.0
        assert body["data"][0]["category"] == "Food"

    def test_upload_excel_serial_dates(self, client):
        content = _excel_serial_dates_wb()
        resp = _upload(client, "dates.xlsx", content)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["valid_rows"] == 1
        assert body["data"][0]["date"] is not None
        assert "2026" in body["data"][0]["date"]

    def test_upload_optional_description_missing(self, client):
        """Row without description should still parse."""
        rows = [("2026-05-01", 100000, "Food", None)]
        content = _sample_wb_bytes(rows=rows)
        resp = _upload(client, "expenses.xlsx", content)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["valid_rows"] == 1

    def test_upload_whitespace_in_headers(self, client):
        """Extra spaces in headers should be normalised."""
        wb = Workbook()
        ws = wb.active
        ws.append(["  Date  ", " Amount "])
        ws.append(["2026-01-01", 1000])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = _upload(client, "spaces.xlsx", buf.read())
        assert resp.status_code == 200, resp.text


class TestUploadValidation:
    """Edge cases that should return 400."""

    def test_upload_no_file(self, client):
        resp = client.post("/api/excel/upload")
        assert resp.status_code == 422  # FastAPI validation

    def test_upload_invalid_extension(self, client):
        resp = _upload(client, "notes.txt", b"hello")
        assert resp.status_code == 400
        assert "format" in resp.json()["detail"].lower()

    def test_upload_empty_file(self, client):
        resp = _upload(client, "empty.xlsx", _empty_wb_bytes())
        assert resp.status_code == 400
        assert "data row" in resp.json()["detail"].lower()

    def test_upload_corrupt_file(self, client):
        resp = _upload(client, "corrupt.xlsx", b"this is not an xlsx file")
        assert resp.status_code == 400

    def test_upload_missing_amount_column(self, client):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Category", "Description"])
        ws.append(["2026-01-01", "Food", "Lunch"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = _upload(client, "noamount.xlsx", buf.read())
        assert resp.status_code == 400
        assert "amount" in resp.json()["detail"].lower()

    def test_upload_missing_date_column(self, client):
        wb = Workbook()
        ws = wb.active
        ws.append(["Amount", "Category"])
        ws.append([5000, "Food"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = _upload(client, "no-date.xlsx", buf.read())
        assert resp.status_code == 400
        assert "date" in resp.json()["detail"].lower()

    def test_upload_oversized_file(self, client, monkeypatch):
        """Patch MAX_FILE_SIZE to 10 bytes and verify rejection."""
        # Patch the value in the *route* namespace — it was imported there at startup.
        monkeypatch.setattr("routes.excel.MAX_FILE_SIZE", 10)
        content = _sample_wb_bytes()
        resp = _upload(client, "big.xlsx", content)
        assert resp.status_code == 400
        assert "too large" in resp.json()["detail"].lower()

    def test_upload_skips_invalid_rows(self, client):
        rows = [
            ("2026-01-01", 1000),
            ("bad-daate-mmd", 2000),
            ("2026-03-01", "not-a-number"),
            ("2026-04-01", 4000),
        ]
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Amount"])
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = _upload(client, "mixed.xlsx", buf.read())
        assert resp.status_code == 200
        body = resp.json()
        assert body["valid_rows"] == 2
        assert body["skipped_count"] == 2
        assert len(body["skipped"]) == 2


# ──────────────────────── Preview tests ────────────────────────────────
class TestPreview:
    def test_preview_single_sheet(self, client):
        content = _sample_wb_bytes()
        resp = _preview(client, "expenses.xlsx", content)
        assert resp.status_code == 200
        body = resp.json()
        assert "sheets" in body
        sheets = body["sheets"]
        assert len(sheets) == 1
        sheet = list(sheets.values())[0]
        assert "headers" in sheet
        assert "sample_rows_cells" in sheet
        assert sheet["total_rows"] == 2

    def test_preview_multi_sheet(self, client):
        content = _multi_sheet_wb()
        resp = _preview(client, "multi.xlsx", content)
        assert resp.status_code == 200
        body = resp.json()
        sheets = body["sheets"]
        assert len(sheets) == 2
        assert "Expenses" in sheets
        assert "Budgets" in sheets
        assert len(sheets["Expenses"]["headers"]) == 2
        assert sheets["Expenses"]["total_rows"] == 2

    def test_preview_csv(self, client):
        csv_content = "A,B,C\r\n1,2,3\r\n4,5,6\r\n"
        resp = _preview(
            client, "simple.csv", csv_content.encode("utf-8")
        )
        assert resp.status_code == 200
        body = resp.json()
        sheets = body["sheets"]
        assert "Sheet1" in sheets
        s = sheets["Sheet1"]
        assert s["headers"] == ["A", "B", "C"]
        assert s["total_rows"] == 2


# ──────────────────────── Export tests ────────────────────────────────
payload = [
    {"date": "2026-01-01", "amount": 1000.0, "category": "Food"},
    {"date": "2026-02-01", "amount": 2000.0, "category": "Rent"},
]

class TestExportXlsx:
    def test_export_xlsx_basic(self, client):
        resp = client.post(
            "/api/excel/export/xlsx",
            json={"data": payload},
        )
        assert resp.status_code == 200
        assert (
            resp.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in resp.headers["content-disposition"]
        assert len(resp.content) > 0
        # Verify the file is a valid xlsx we can re-open
        wb = openpyxl.load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(1, 1).value == "date"
        assert ws.cell(2, 1).value == "2026-01-01"
        wb.close()

    def test_export_xlsx_custom_sheet_name(self, client):
        resp = client.post(
            "/api/excel/export/xlsx",
            json={"data": payload, "sheet_name": "MyExport"},
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(filename=BytesIO(resp.content))
        assert wb.sheetnames == ["MyExport"]
        wb.close()

    def test_export_xlsx_custom_columns(self, client):
        resp = client.post(
            "/api/excel/export/xlsx",
            json={
                "data": payload,
                "columns": ["date", "amount"],
            },
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(1, 1).value == "date"
        assert ws.cell(1, 2).value == "amount"
        assert ws.cell(1, 3).value is None  # no category column
        wb.close()

    def test_export_xlsx_empty_data(self, client):
        resp = client.post(
            "/api/excel/export/xlsx",
            json={"data": []},
        )
        assert resp.status_code == 400


class TestExportCsv:
    def test_export_csv_basic(self, client):
        resp = client.post(
            "/api/excel/export/csv",
            json={"data": payload},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")
        assert "attachment" in resp.headers["content-disposition"]
        text = resp.content.decode("utf-8-sig")
        assert "date,amount,category" in text
        assert "2026-01-01,1000.0,Food" in text

    def test_export_csv_empty_data(self, client):
        resp = client.post(
            "/api/excel/export/csv",
            json={"data": []},
        )
        assert resp.status_code == 400


# ──────────────────────── Route registration ──────────────────────────
def test_excel_routes_registered(registered_paths):
    """All four Excel endpoints must appear in the registered routes."""
    assert "/api/excel/upload" in registered_paths
    assert "/api/excel/preview" in registered_paths
    assert "/api/excel/export/xlsx" in registered_paths
    assert "/api/excel/export/csv" in registered_paths