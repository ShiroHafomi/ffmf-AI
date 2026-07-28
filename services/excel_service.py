"""Excel file parsing, extraction, and generation — temp processing only, no DB.

Mirrors the column-name aliasing and date-parsing logic from the Node
backend (``backend/src/routes/expenses.ts``) so the Python AI microservice
can independently accept and produce the same file formats.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger("ffms")

# ─────────────────────────── Constants ───────────────────────────
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB (matches Node multer config)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv"}

# ── Column-name aliases (English / Vietnamese / case-insensitive) ──
# Mirrors ``COLUMN_ALIASES`` in backend/src/routes/expenses.ts
COLUMN_ALIASES: dict[str, str] = {
    # English
    "date": "date",
    "expense_date": "date",
    "expense date": "date",
    "amount": "amount",
    "total": "amount",
    "cost": "amount",
    "price": "amount",
    "sum": "amount",
    "value": "amount",
    "category": "category",
    "category_name": "category",
    "category name": "category",
    "cat": "category",
    "description": "description",
    "desc": "description",
    "note": "description",
    "notes": "description",
    "detail": "description",
    "details": "description",
    "memo": "description",
    # Vietnamese
    "ngày": "date",
    "ngay": "date",
    "số tiền": "amount",
    "so tien": "amount",
    "tiền": "amount",
    "tien": "amount",
    "thành tiền": "amount",
    "thanh tien": "amount",
    "danh mục": "category",
    "danh muc": "category",
    "phân loại": "category",
    "phan loai": "category",
    "loại": "category",
    "loai": "category",
    "mô tả": "description",
    "mo ta": "description",
    "diễn giải": "description",
    "dien giai": "description",
    "ghi chú": "description",
    "ghi chu": "description",
}


# ──────────────────────── Header normalisation ────────────────────────
def normalise_header(header: str) -> str:
    """Normalise a column header and return the canonical name.

    Falls back to the original, stripped-and-lowercased value if no alias
    matches, so callers can still inspect unknown columns.
    """
    key = header.strip().lower()
    # Collapse multiple spaces (e.g. two spaces between words)
    key = " ".join(key.split())
    return COLUMN_ALIASES.get(key, key)


# ──────────────────────── Date parsing ────────────────────────────────
def parse_excel_date(raw: Any) -> str | None:
    """Parse a date value into ``YYYY-MM-DD``.

    Handles
    ------
    * Excel serial date numbers (int / float, epoch 1899-12-30)
    * ``YYYY-MM-DD`` strings
    * ``DD/MM/YYYY`` strings (VN/UK preferred in ambiguous cases)
    * Python ``datetime`` objects passed through by the reader library
    """
    if raw is None or raw == "":
        return None

    # openpyxl may return a Python datetime directly
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")

    # Excel serial date (number) — epoch is Dec 30 1899 (1900 leap-year bug)
    if isinstance(raw, (int, float)):
        try:
            excel_epoch = datetime(1899, 12, 30)
            d = excel_epoch + timedelta(days=float(raw))
            if d.year >= 1900:
                return d.strftime("%Y-%m-%d")
        except (OverflowError, OSError, ValueError):
            pass
        return None

    s = str(raw).strip()
    if not s:
        return None

    # YYYY-MM-DD
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        parts = s[:10].split("-")
        if len(parts) == 3 and parts[0].isdigit():
            y, mo, d_ = parts
            return f"{int(y):04d}-{int(mo):02d}-{int(d_):02d}"

    # DD/MM/YYYY or MM/DD/YYYY
    slash_parts = s.split("/")
    if len(slash_parts) == 3:
        a, b, y = slash_parts[0].strip(), slash_parts[1].strip(), slash_parts[2].strip()
        if not (a.isdigit() and b.isdigit() and y.isdigit()):
            pass
        else:
            ai, bi = int(a), int(b)
            if ai > 12 and bi <= 12:
                # a must be day
                return f"{int(y):04d}-{bi:02d}-{ai:02d}"
            if bi > 12 and ai <= 12:
                # b must be day
                return f"{int(y):04d}-{ai:02d}-{bi:02d}"
            if ai <= 12 and bi <= 12:
                # Ambiguous: assume DD/MM/YYYY (VN locale)
                return f"{int(y):04d}-{bi:02d}-{ai:02d}"

    # Last resort: generic JS-style date parse
    try:
        from dateutil.parser import parse as dateutil_parse

        return dateutil_parse(s).strftime("%Y-%m-%d")
    except Exception:
        pass

    # Fallback: basic Python date parse (handles "January 1, 2026" etc.)
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%d %b %Y",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


# ──────────────────────── Amount parsing ──────────────────────────────
def parse_amount(raw: Any) -> float | None:
    """Parse a numeric / currency cell into a float.

    Removes currency symbols (₫, $, €, £), collapses whitespace, and
    handles both Western (``1,234.56``) and Vietnamese (``1.234,56``)
    thousand / decimal conventions.
    """
    if raw is None or raw == "":
        return None

    # Already a number
    if isinstance(raw, (int, float)):
        return float(raw)

    s = str(raw).strip()
    if not s:
        return None

    # Strip common currency symbols / suffixes
    for symbol in ("₫", "VND", "$", "€", "£", "¥", "đ", "vnd"):
        s = s.replace(symbol, "")
    s = s.strip()

    if not s:
        return None

    # Detect Vietnamese-style formatting: "1.234.567,89" or "1.234,56"
    # If there's a comma, it's likely the decimal separator in VN format
    if "," in s and "." in s:
        # Which comes last? Last = decimal separator
        last_dot = s.rfind(".")
        last_comma = s.rfind(",")
        if last_comma > last_dot:
            # 12.345,67 => comma is decimal
            s = s.replace(".", "").replace(",", ".")
        else:
            # 12,345.67 => dot is decimal
            s = s.replace(",", "")
    elif "," in s:
        # "1.234,56" style: if only comma present and 1-3 digits after last comma
        # assume comma is decimal
        last_comma = s.rfind(",")
        digits_after = len(s) - last_comma - 1
        if digits_after <= 3 and last_comma > 0:
            # "1,56" or "1234,56": comma is decimal
            s = s.replace(",", ".")
        elif digits_after == 3 and s.count(",") == 1:
            # "1234,560" → ambiguous
            s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        pass

    return None


# ──────────────────────── File parsing ────────────────────────────────
def parse_file(content: bytes, filename: str) -> list[list[Any]]:
    """Read a file into an in-memory 2D list of cell values.

    Returns every raw cell value. The first row is assumed to be a header
    by upstream callers.

    Raises :exc:`ValueError` when the file is corrupt, unsupported, or empty.
    """
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError("Unsupported file format.")

    try:
        if ext == "csv":
            return _parse_csv(content)
        if ext == "xlsx":
            return _parse_xlsx(content)
        # xls (legacy binary)
        return _parse_xls(content)
    except ValueError:
        raise
    except Exception as e:
        logging.getLogger("ffms").warning(
            "File parsing failed: %s", e, exc_info=True
        )
        raise ValueError(
            "Could not read the file. It may be corrupted or in an unsupported format."
        ) from e


def _parse_csv(content: bytes) -> list[list[Any]]:
    """Parse CSV bytes into list[list[cell]]."""
    text = content.decode("utf-8-sig")  # handle BOM
    reader = csv.reader(StringIO(text))
    rows = [list(row) for row in reader]
    if not rows:
        raise ValueError("File appears to be empty.")
    return rows


def _parse_xlsx(content: bytes) -> list[list[Any]]:
    """Parse .xlsx bytes via openpyxl into list[list[cell]]."""
    wb = openpyxl.load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active sheet.")
        rows: list[list[Any]] = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows.append(list(row))
        if not rows:
            raise ValueError("File appears to be empty.")
        return rows
    finally:
        wb.close()


def _parse_xls(content: bytes) -> list[list[Any]]:
    """Parse legacy .xls bytes via xlrd into list[list[cell]]."""
    import xlrd  # lazy import — only needed for legacy .xls

    wb = xlrd.open_workbook(file_contents=content)
    if wb.nsheets == 0:
        raise ValueError("Workbook has no sheets.")
    ws = wb.sheet_by_index(0)
    rows: list[list[Any]] = []
    for r in range(ws.nrows):
        row_vals: list[Any] = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
            if cell.ctype == 3:  # XL_CELL_DATE
                try:
                    row_vals.append(
                        datetime(
                            *xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        )
                    )
                except Exception:
                    row_vals.append(cell.value)
            elif cell.ctype == 5:  # XL_CELL_ERROR
                row_vals.append(None)
            else:
                row_vals.append(cell.value)
        rows.append(row_vals)
    if not rows:
        raise ValueError("File appears to be empty.")
    return rows


# ────────────────── Structured expense extraction ─────────────────────
def extract_structured_data(rows: list[list[Any]]) -> dict[str, Any]:
    """Detect expense columns via header-row aliases and build structured rows.

    Returns
    -------
    dict with keys: headers, normalised_headers, data, skipped,
                    total_rows, valid_rows, skipped_count
    """
    if len(rows) < 2:
        raise ValueError(
            "The file must have a header row and at least one data row."
        )

    headers = [str(c) if c is not None else "" for c in rows[0]]
    normalised = [normalise_header(h) for h in headers]

    # --- Column mapping ---
    col_map: dict[str, int | None] = {
        "date": None,
        "amount": None,
        "category": None,
        "description": None,
    }
    for i, nh in enumerate(normalised):
        if nh in col_map and col_map[nh] is None:
            col_map[nh] = i  # take the first match

    if col_map["date"] is None or col_map["amount"] is None:
        missing = []
        if col_map["date"] is None:
            missing.append("Date")
        if col_map["amount"] is None:
            missing.append("Amount")
        raise ValueError(
            f"Required column(s) not found: {', '.join(missing)}. "
            "The file needs at least a 'Date' column (ngay) and an 'Amount' column."
        )

    data: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for idx, row in enumerate(rows[1:], start=2):  # row 2 = first data row
        date_val = parse_excel_date(_cell(row, col_map["date"]))
        amount_val = parse_amount(_cell(row, col_map["amount"]))

        if date_val is None or amount_val is None:
            skipped.append(
                {"row": idx, "reason": "Missing or invalid date / amount."}
            )
            continue

        record: dict[str, Any] = {"date": date_val, "amount": amount_val}
        if col_map["category"] is not None:
            cat = _cell(row, col_map["category"])
            record["category"] = str(cat).strip() if cat else ""
        if col_map["description"] is not None:
            desc = _cell(row, col_map["description"])
            record["description"] = str(desc).strip() if desc else ""
        data.append(record)

    return {
        "headers": [_str_cell(c) for c in headers],
        "normalised_headers": normalised,
        "data": data,
        "skipped": skipped,
        "total_rows": len(rows) - 1,
        "valid_rows": len(data),
        "skipped_count": len(skipped),
    }


def _cell(row: list[Any], idx: int | None) -> Any:
    """Safely read a cell value; return None if index is out of bounds."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _str_cell(cell: Any) -> str:
    """Convert a cell value to a trimmed string."""
    if cell is None:
        return ""
    return str(cell).strip()


# ─────────────── General-purpose multi-sheet preview ─────────────────
def extract_all_sheets(content: bytes, filename: str) -> dict[str, dict[str, Any]]:
    """Return every sheet with headers and up to 5 sample rows."""
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()

    sheets: dict[str, dict[str, Any]] = {}

    if ext == "csv":
        rows = _parse_csv(content)
        headers = [str(c) if c is not None else "" for c in (rows[0] if rows else [])]
        sample = rows[1:6] if len(rows) > 1 else []
        sheets["Sheet1"] = {
            "headers": [_str_cell(h) for h in headers],
            "sample_rows_cells": sample,
            "total_rows": len(rows) - 1 if rows else 0,
        }
        return sheets

    if ext == "xlsx":
        wb = openpyxl.load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        try:
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                rows = []
                for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
                    rows.append(list(row))
                # Count total rows (scan full sheet)
                total = 0
                for _ in ws.iter_rows(min_row=1, values_only=True):
                    total += 1
                headers = [str(c) if c is not None else "" for c in (rows[0] if rows else [])]
                sample = rows[1:] if len(rows) > 1 else []
                sheets[str(ws_name)] = {
                    "headers": [_str_cell(h) for h in headers],
                    "sample_rows_cells": sample,
                    "total_rows": max(0, total - 1),
                }
        finally:
            wb.close()
        return sheets

    # xls — xlrd does not support read_only, just load everything
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    for i in range(wb.nsheets):
        ws = wb.sheet_by_index(i)
        headers = []
        if ws.nrows > 0:
            headers = [_str_cell(ws.cell_value(0, c)) for c in range(ws.ncols)]
        sample: list[list[Any]] = []
        for r in range(1, min(ws.nrows, 6)):
            sample.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        sheets[str(ws.name)] = {
            "headers": headers,
            "sample_rows_cells": sample,
            "total_rows": max(0, ws.nrows - 1),
        }
    return sheets


# ──────────────────────── Export generation ───────────────────────────
HEADER_FILL = PatternFill(start_color="6C8CFF", end_color="6C8CFF", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=11)
AMOUNT_FORMAT = '#,##0.00'


def generate_excel_bytes(
    data: list[dict[str, Any]],
    columns: list[str] | None = None,
    sheet_name: str = "Sheet1",
) -> bytes:
    """Build an .xlsx workbook and return its bytes (fully in-memory)."""
    if columns is None and data:
        columns = list(data[0].keys())
    if not columns:
        raise ValueError("Cannot generate Excel file: no columns specified.")

    wb = openpyxl.Workbook()
    ws = wb.active

    # Avoid openpyxl 3.0+ warning about active sheet being "Sheet"
    if ws is not None:
        ws.title = sheet_name or "Sheet1"

        # Header row
        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for ri, record in enumerate(data, 2):
            for ci, col_name in enumerate(columns, 1):
                val = record.get(col_name, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = DATA_FONT
                # Format numeric columns
                if isinstance(val, float):
                    cell.number_format = AMOUNT_FORMAT
                elif isinstance(val, int):
                    cell.number_format = "#,##0"

        # Auto-fit column widths
        for ci, col_name in enumerate(columns, 1):
            max_width = len(str(col_name)) + 2
            for ri in range(2, len(data) + 2):
                cell = ws.cell(row=ri, column=ci)
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_width:
                    max_width = cell_len
            ws.column_dimensions[get_column_letter(ci)].width = min(max_width + 2, 40)

        # Auto-filter on header range
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data)+1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_csv_string(
    data: list[dict[str, Any]],
    columns: list[str] | None = None,
) -> str:
    """Build a CSV string from a list of dicts (fully in-memory)."""
    if columns is None and data:
        columns = list(data[0].keys())
    if not columns:
        columns = []

    out = StringIO()
    writer = csv.DictWriter(out, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for record in data:
        # Convert every value to a string-safe form
        row: dict[str, Any] = {}
        for k in columns:
            v = record.get(k, "")
            if isinstance(v, (int, float)):
                row[k] = v
            elif v is None:
                row[k] = ""
            else:
                row[k] = str(v)
        writer.writerow(row)

    return out.getvalue()